import requests
import pandas as pd
import time
import math
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


class AccuratePollutionSourceCollector:
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://restapi.amap.com/v3/place/around"

        # 新增：地理编码API地址，用于获取地址详细信息
        self.geocode_url = "https://restapi.amap.com/v3/geocode/regeo"

        # API调用计数器
        self.api_call_count = 0

        # 按照要求定义的九大类污染源企业类型
        self.pollution_industry_types = {
            # 1. 金属矿采选业
            "金属矿采选业": [
                "金属矿", "矿山", "采矿", "选矿", "采选", "矿业", "矿场", "矿区",
                "铁矿", "铜矿", "铅矿", "锌矿", "铝矿", "金矿", "银矿", "镍矿",
                "锡矿", "钨矿", "钼矿", "锰矿", "铬矿", "钴矿", "锑矿", "汞矿",
                "稀有金属矿", "稀土矿", "贵金属矿", "黑色金属矿", "有色金属矿",
                "采石场", "砂石矿", "石英矿", "云母矿", "长石矿", "高岭土矿"
            ],

            # 2. 石油加工业
            "石油加工业": [
                "石油加工", "炼油", "石油炼制", "石油化工", "石化", "炼化",
                "催化裂化", "加氢裂化", "催化重整", "延迟焦化", "烷基化",
                "润滑油", "沥青", "石油焦", "石油制品", "成品油", "原油加工",
                "燃料油", "石脑油", "炼油厂", "石化厂"
            ],

            # 3. 炼焦业
            "炼焦业": [
                "炼焦", "焦化", "焦炭", "焦炉", "煤焦化", "焦化厂",
                "煤气", "焦炉煤气", "煤焦油", "焦化产品", "焦化车间"
            ],

            # 4. 化学原料和化学制品制造业
            "化学原料和化学制品制造业": [
                "化工厂", "化学原料", "化学制品", "基础化工", "无机化工", "有机化工",
                "硫酸", "盐酸", "硝酸", "磷酸", "氢氟酸", "烧碱", "纯碱",
                "氯碱", "化肥", "农药", "涂料", "油漆", "颜料", "染料",
                "油墨", "合成材料", "塑料", "橡胶", "化纤", "树脂",
                "助剂", "添加剂", "催化剂", "溶剂", "试剂", "专用化学品",
                "电子化学品", "医药中间体", "表面活性剂", "水处理剂"
            ],

            # 5. 医药制造业
            "医药制造业": [
                "制药", "药厂", "医药", "制药厂", "药品", "原料药",
                "化学制药", "生物制药", "中药", "中药提取", "医药中间体",
                "制剂", "抗生素", "维生素", "激素", "疫苗", "血液制品",
                "药用辅料", "医药原料", "药品生产"
            ],

            # 6. 黑色金属冶炼和压延加工业
            "黑色金属冶炼和压延加工业": [
                "钢铁", "炼铁", "炼钢", "轧钢", "铁合金", "不锈钢",
                "钢铁厂", "炼铁厂", "炼钢厂", "轧钢厂", "冷轧", "热轧",
                "钢板", "钢管", "型钢", "线材", "螺纹钢", "钢铁冶炼",
                "钢铁加工", "金属压延", "钢材加工"
            ],

            # 7. 有色金属冶炼和压延加工业
            "有色金属冶炼和压延加工业": [
                "有色金属", "铜冶炼", "铝冶炼", "铅冶炼", "锌冶炼", "镍冶炼",
                "锡冶炼", "锑冶炼", "镁冶炼", "钛冶炼", "稀有金属冶炼",
                "电解铝", "电解铜", "铜加工", "铝加工", "铜材", "铝材",
                "有色金属压延", "有色金属加工", "铜箔", "铝箔"
            ],

            # 8. 电池制造业
            "电池制造业": [
                "电池", "蓄电池", "锂电池", "铅酸电池", "镍氢电池", "镍镉电池",
                "锂离子电池", "电池厂", "电池生产", "电芯", "电池组装",
                "电池材料", "正极材料", "负极材料", "电解液", "隔膜",
                "电池制造", "电池研发", "电池加工"
            ],

            # 9. 固体废物处理业
            "固体废物处理业": [
                "固体废物", "危废", "危险废物", "医疗废物", "工业固废",
                "生活垃圾", "废物处理", "废物处置", "垃圾处理", "垃圾焚烧",
                "填埋场", "废物利用", "资源回收", "废物回收",
                "废金属回收", "电子废物", "废电池回收", "废油处理",
                "废酸处理", "废物填埋", "废物焚烧", "固废处理", "危废处置"
            ]
        }

    def increment_api_call(self):
        """增加API调用计数"""
        self.api_call_count += 1
        return self.api_call_count

    # ==================== 坐标转换函数 ====================
    def wgs84_to_gcj02(self, lng, lat):
        """
        WGS84转GCJ02坐标系（火星坐标系）
        使用公开的转换算法
        """
        pi = 3.1415926535897932384626
        a = 6378245.0
        ee = 0.00669342162296594323

        def transform_lng(lng, lat):
            ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + 0.1 * lng * lat + 0.1 * math.sqrt(abs(lng))
            ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 * math.sin(2.0 * lng * pi)) * 2.0 / 3.0
            ret += (20.0 * math.sin(lng * pi) + 40.0 * math.sin(lng / 3.0 * pi)) * 2.0 / 3.0
            ret += (150.0 * math.sin(lng / 12.0 * pi) + 300.0 * math.sin(lng / 30.0 * pi)) * 2.0 / 3.0
            return ret

        def transform_lat(lng, lat):
            ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + 0.1 * lng * lat + 0.2 * math.sqrt(abs(lng))
            ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 * math.sin(2.0 * lng * pi)) * 2.0 / 3.0
            ret += (20.0 * math.sin(lat * pi) + 40.0 * math.sin(lat / 3.0 * pi)) * 2.0 / 3.0
            ret += (160.0 * math.sin(lat / 12.0 * pi) + 320.0 * math.sin(lat * pi / 30.0)) * 2.0 / 3.0
            return ret

        dlat = transform_lat(lng - 105.0, lat - 35.0)
        dlng = transform_lng(lng - 105.0, lat - 35.0)

        radlat = lat / 180.0 * pi
        magic = math.sin(radlat)
        magic = 1 - ee * magic * magic
        sqrtmagic = math.sqrt(magic)

        dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
        dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)

        mglat = lat + dlat
        mglng = lng + dlng

        return mglng, mglat

    def gcj02_to_wgs84(self, lng, lat):
        """
        GCJ02转WGS84坐标系（火星坐标系转地球坐标系）
        使用迭代法进行反向转换
        """
        # 初始值设为GCJ02坐标
        wgs_lng, wgs_lat = lng, lat

        # 迭代计算，直到误差小于1e-6度（约0.1米）
        for i in range(5):  # 一般3-5次迭代即可达到足够精度
            gcj_lng, gcj_lat = self.wgs84_to_gcj02(wgs_lng, wgs_lat)
            delta_lng = gcj_lng - lng
            delta_lat = gcj_lat - lat
            wgs_lng -= delta_lng
            wgs_lat -= delta_lat

            # 如果误差已经很小，提前退出
            if abs(delta_lng) < 1e-7 and abs(delta_lat) < 1e-7:
                break

        return wgs_lng, wgs_lat

    # ==================== 坐标转换函数结束 ====================

    def calculate_distance(self, lng1, lat1, lng2, lat2):
        """使用Haversine公式计算两点间距离（米）"""
        lng1, lat1, lng2, lat2 = map(math.radians, [lng1, lat1, lng2, lat2])

        dlng = lng2 - lng1
        dlat = lat2 - lat1
        a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
        c = 2 * math.asin(math.sqrt(a))
        r = 6371
        return c * r * 1000

    # ==================== 新增：获取地址详细信息函数 ====================
    def get_address_details(self, lng, lat):
        """
        通过逆地理编码API获取地址详细信息
        返回省份、城市、区县等地址信息
        """
        try:
            # 调用逆地理编码API
            params = {
                'key': self.api_key,
                'location': f'{lng},{lat}',
                'output': 'json',
                'extensions': 'all'
            }

            # 增加API调用计数
            self.increment_api_call()

            response = requests.get(self.geocode_url, params=params, timeout=5)
            data = response.json()

            if data.get('status') == '1' and data.get('regeocode'):
                address_component = data['regeocode']['addressComponent']

                # 提取地址信息
                address_details = {
                    'province': address_component.get('province', ''),
                    'city': address_component.get('city', ''),
                    'district': address_component.get('district', ''),
                    'township': address_component.get('township', ''),
                    'formatted_address': data['regeocode'].get('formatted_address', '')
                }

                # 有些情况下city为空，使用province作为city
                if not address_details['city'] and address_details['province']:
                    address_details['city'] = address_details['province']

                return address_details
            else:
                return {
                    'province': '',
                    'city': '',
                    'district': '',
                    'township': '',
                    'formatted_address': ''
                }

        except Exception as e:
            print(f"  获取地址信息出错: {e}")
            return {
                'province': '',
                'city': '',
                'district': '',
                'township': '',
                'formatted_address': ''
            }

    # ==================== 地址信息函数结束 ====================

    def is_pollution_enterprise(self, name, enterprise_type):
        """判断是否为污染源企业"""
        combined_text = (name + ' ' + enterprise_type).lower()

        for industry, keywords in self.pollution_industry_types.items():
            for keyword in keywords:
                if keyword.lower() in combined_text:
                    return True, industry

        return False, "其他"

    # ==================== 修改搜索函数 ====================
    def search_nearby_pollution_sources(self, lng, lat, radius=5000):
        """搜索周边污染源企业"""
        # 1. 采样点坐标转换：WGS84转GCJ02
        gcj_lng, gcj_lat = self.wgs84_to_gcj02(lng, lat)

        print(f"  采样点坐标转换完成 (WGS84→GCJ02)，开始搜索5km范围内污染源...")

        all_enterprises = []

        type_codes = [
            "060200",  # 工厂 - 最直接相关的分类
            "060300",  # 园区 - 工业园区可能包含多个污染源
            "170000",  # 公司企业（制造业） - 制造业相关
            "060100",  # 公司 - 可能包含化工、医药等公司
            "010300",  # 汽车维修 - 可能产生废油
            "060400",  # 农林牧渔基地 - 可能涉及农药化肥
            "130000",  # 基础设施 - 可能包含污水处理厂等
        ]

        for type_code in type_codes:
            page = 1
            while True:
                # 修复：确保try-except结构完整
                try:
                    params = {
                        'key': self.api_key,
                        'location': f'{gcj_lng},{gcj_lat}',
                        'radius': radius,
                        'types': type_code,
                        'offset': 50,
                        'page': page,
                        'extensions': 'all'
                    }

                    # 增加API调用计数
                    self.increment_api_call()
                    if self.api_call_count % 50 == 0:
                        print(f"  API调用次数: {self.api_call_count}次")

                    response = requests.get(self.base_url, params=params, timeout=10)
                    data = response.json()

                    # 检查API密钥是否有效
                    if data.get('status') == '0':
                        error_info = data.get('info', '')
                        if 'key' in error_info.lower() or '配额' in error_info:
                            raise Exception(f"API密钥无效或配额已用完: {error_info}")

                    if data['status'] == '1' and int(data.get('count', 0)) > 0:
                        pois = data['pois']

                        for poi in pois:
                            if 'location' in poi and poi['location']:
                                try:
                                    # 2. 获取企业GCJ02坐标
                                    enterprise_gcj_lng, enterprise_gcj_lat = map(float, poi['location'].split(','))

                                    # 3. 企业坐标转换：GCJ02转WGS84
                                    enterprise_wgs_lng, enterprise_wgs_lat = self.gcj02_to_wgs84(
                                        enterprise_gcj_lng, enterprise_gcj_lat
                                    )

                                    # 4. 使用WGS84坐标计算距离
                                    distance = self.calculate_distance(lng, lat, enterprise_wgs_lng, enterprise_wgs_lat)

                                    # 判断是否为污染源企业
                                    is_pollution, industry_type = self.is_pollution_enterprise(
                                        poi.get('name', ''),
                                        poi.get('type', '')
                                    )

                                    if is_pollution:
                                        # 5. 新增：获取企业的详细地址信息
                                        address_details = self.get_address_details(
                                            enterprise_gcj_lng, enterprise_gcj_lat
                                        )

                                        enterprise_info = {
                                            'enterprise_name': poi.get('name', ''),
                                            'enterprise_type': poi.get('type', ''),
                                            'enterprise_address': poi.get('address', ''),

                                            # 新增地址信息字段
                                            'province': address_details['province'],
                                            'city': address_details['city'],
                                            'district': address_details['district'],
                                            'township': address_details['township'],
                                            'formatted_address': address_details['formatted_address'],

                                            # 坐标信息：输出WGS84坐标
                                            'enterprise_longitude_wgs84': round(enterprise_wgs_lng, 6),
                                            'enterprise_latitude_wgs84': round(enterprise_wgs_lat, 6),

                                            # 保留原始GCJ02坐标以供参考
                                            'enterprise_longitude_gcj02': round(enterprise_gcj_lng, 6),
                                            'enterprise_latitude_gcj02': round(enterprise_gcj_lat, 6),

                                            'distance_to_sample': round(distance),
                                            'pollution_industry': industry_type,
                                            'business_area': poi.get('business_area', ''),
                                            'telephone': poi.get('tel', ''),
                                            'poi_id': poi.get('id', '')
                                        }
                                        all_enterprises.append(enterprise_info)
                                        print(f"    发现: {enterprise_info['enterprise_name']} [{industry_type}]")

                        # 修复：这个if语句现在在try块内部，是正确的
                        if len(pois) < 50:
                            break
                        page += 1

                        time.sleep(0.2)
                    else:
                        break

                except Exception as e:
                    print(f"  API请求错误: {e}")
                    if 'key' in str(e).lower() or '配额' in str(e):
                        raise e
                    break

        return all_enterprises
    # ==================== 搜索函数修改结束 ====================


def load_sample_data(csv_file_path):
    """从CSV文件加载采样点数据"""
    try:
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
        sample_data = None

        for encoding in encodings:
            try:
                sample_data = pd.read_csv(csv_file_path, encoding=encoding)
                print(f"成功使用 {encoding} 编码读取文件")
                break
            except UnicodeDecodeError:
                continue

        if sample_data is None:
            print("无法读取CSV文件，请检查文件路径和编码")
            return None

        required_columns = ['序号', '经度', '纬度']
        missing_columns = [col for col in required_columns if col not in sample_data.columns]

        if missing_columns:
            print(f"CSV文件缺少必要的列: {missing_columns}")
            print(f"文件中的列: {list(sample_data.columns)}")
            return None

        sample_data = sample_data.rename(columns={
            '序号': 'sample_id',
            '经度': 'longitude',
            '纬度': 'latitude'
        })

        print(f"成功加载 {len(sample_data)} 个采样点")
        return sample_data

    except Exception as e:
        print(f"加载CSV文件时出错: {e}")
        return None


class ExcelWriter:
    """Excel写入器，支持实时追加数据"""

    def __init__(self, output_file):
        self.output_file = output_file
        self.initialized = False

    def initialize_excel(self, columns):
        """初始化Excel文件"""
        df = pd.DataFrame(columns=columns)
        df.to_excel(self.output_file, index=False)
        self.initialized = True
        print(f"已创建Excel文件: {self.output_file}")

    def append_to_excel(self, data_list):
        """追加数据到Excel文件"""
        if not data_list:
            return

        if not self.initialized:
            columns = list(data_list[0].keys())
            self.initialize_excel(columns)

        try:
            if os.path.exists(self.output_file):
                existing_df = pd.read_excel(self.output_file)
            else:
                existing_df = pd.DataFrame()

            new_df = pd.DataFrame(data_list)

            if not existing_df.empty:
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            else:
                combined_df = new_df

            combined_df.to_excel(self.output_file, index=False)

            print(f"  已写入 {len(data_list)} 条数据到Excel，当前总计 {len(combined_df)} 条")

        except Exception as e:
            print(f"  写入Excel时出错: {e}")

    def format_excel(self):
        """格式化Excel文件"""
        try:
            if not os.path.exists(self.output_file):
                return

            workbook = load_workbook(self.output_file)
            worksheet = workbook.active

            header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 调整列宽以适应新字段
            column_widths = {
                'A': 12, 'B': 15, 'C': 15,  # 采样点信息
                'D': 30, 'E': 30, 'F': 40,  # 企业基本信息
                'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 40,  # 地址信息
                'L': 15, 'M': 15, 'N': 15, 'O': 15,  # 坐标信息
                'P': 15, 'Q': 20, 'R': 15, 'S': 20  # 其他信息
            }

            for col, width in column_widths.items():
                if col in worksheet.column_dimensions:
                    worksheet.column_dimensions[col].width = width

            workbook.save(self.output_file)
            print(f"已格式化Excel文件: {self.output_file}")

        except Exception as e:
            print(f"格式化Excel时出错: {e}")


def collect_pollution_sources(sample_data, api_key, output_file):
    """
    为每个采样点收集周边污染源企业信息，实时写入Excel
    """
    collector = AccuratePollutionSourceCollector(api_key)
    excel_writer = ExcelWriter(output_file)

    total_sources_found = 0
    processed_samples = 0

    print("开始收集污染源企业信息...")
    print("坐标转换说明:")
    print("  - 采样点坐标: WGS84 → GCJ02 (用于API搜索)")
    print("  - 企业坐标: GCJ02 → WGS84 (用于输出结果)")

    for index, sample in sample_data.iterrows():
        processed_samples += 1
        print(f"处理采样点 {processed_samples}/{len(sample_data)}: {sample['sample_id']}")
        print(f"  采样点坐标: {sample['longitude']}, {sample['latitude']} (WGS84)")

        try:
            # 搜索5km范围内的污染源企业
            enterprises = collector.search_nearby_pollution_sources(
                sample['longitude'],
                sample['latitude'],
                radius=5000  # 5km范围
            )

            if enterprises:
                data_to_write = []
                for enterprise in enterprises:
                    result = {
                        'sample_id': sample['sample_id'],
                        'sample_longitude': sample['longitude'],
                        'sample_latitude': sample['latitude'],
                        **enterprise
                    }
                    data_to_write.append(result)

                excel_writer.append_to_excel(data_to_write)
                total_sources_found += len(enterprises)
                print(f"  找到 {len(enterprises)} 个潜在污染源，已实时保存")
            else:
                print(f"  5km范围内未找到污染源")

        except Exception as e:
            error_msg = str(e)
            if 'key' in error_msg.lower() or '配额' in error_msg:
                print(f"\nAPI密钥错误或配额已用完: {error_msg}")
                print("正在保存已收集的数据...")

                excel_writer.format_excel()

                if total_sources_found > 0:
                    print(f"\n已成功收集 {total_sources_found} 个污染源信息")
                    print(f"已处理 {processed_samples}/{len(sample_data)} 个采样点")
                    print(f"数据已保存到: {output_file}")
                else:
                    print("未收集到任何污染源信息")

                return False, total_sources_found, processed_samples, collector.api_call_count
            else:
                print(f"  处理采样点 {sample['sample_id']} 时出错: {e}")

        time.sleep(0.5)  # 避免请求过于频繁

    excel_writer.format_excel()
    return True, total_sources_found, processed_samples, collector.api_call_count


def main():
    # 您的高德地图API密钥
    API_KEY = "e255d6648e4d361f1988cd643988dcbd"  # 请替换为您自己的API密钥

    # CSV文件路径
    csv_file_path = "China2.csv"

    # 输出文件路径
    output_file = '污染源企业信息12.9.xlsx'

    # 加载采样点数据
    sample_data = load_sample_data(csv_file_path)
    if sample_data is None:
        return

    print("\n开始执行污染源企业信息收集...")
    print(f"结果将实时保存到: {output_file}")
    print("搜索范围: 每个采样点周边5km")
    print("坐标系统: 采样点WGS84 → GCJ02 → 企业WGS84")

    # 收集污染源企业信息（实时写入Excel）
    success, total_sources, processed_samples, api_calls = collect_pollution_sources(
        sample_data, API_KEY, output_file
    )

    if success:
        print(f"\n收集完成！")
        print(f"处理采样点数量: {processed_samples}/{len(sample_data)}")
        print(f"找到污染源总数: {total_sources}")
        print(f"API总调用次数: {api_calls}次")
        print(f"结果已保存至: {output_file}")

        # 如果收集到了数据，显示统计信息
        if total_sources > 0:
            try:
                df = pd.read_excel(output_file)
                industry_stats = df['pollution_industry'].value_counts()
                print("\n污染源行业分布:")
                for industry, count in industry_stats.items():
                    print(f"  {industry}: {count}个")

            except Exception as e:
                print(f"读取统计信息时出错: {e}")
    else:
        print(f"\n收集过程中断！")
        print(f"已处理采样点: {processed_samples}/{len(sample_data)}")
        print(f"已找到污染源: {total_sources}个")
        print(f"API调用次数: {api_calls}次")
        print(f"部分结果已保存至: {output_file}")
        print("请检查API密钥或等待配额重置后继续执行")


if __name__ == "__main__":
    main()