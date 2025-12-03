import requests
import pandas as pd
import time
import math
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


class AccuratePollutionSourceCollector:
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://restapi.amap.com/v3/place/around"

        # 全面定义污染源企业类型
        self.pollution_industry_types = {
            # 扩展后的污染源企业类型定义
            "金属矿采选业": [
                "金属矿", "矿山", "采矿", "选矿", "采选", "矿业", "矿场", "矿区",
                "铁矿", "铜矿", "铅矿", "锌矿", "铝矿", "金矿", "银矿", "镍矿",
                "锡矿", "钨矿", "钼矿", "锰矿", "铬矿", "钴矿", "锑矿", "汞矿",
                "稀有金属矿", "稀土矿", "贵金属矿", "黑色金属矿", "有色金属矿",
                "采石场", "砂石矿", "石英矿", "云母矿", "长石矿", "高岭土矿"
            ],

            "金属冶炼加工业": [
                "冶炼", "熔炼", "精炼", "电解", "熔铸", "铸造", "压延", "轧制",
                "炼铁", "炼钢", "钢铁", "轧钢", "铸钢", "不锈钢", "特种钢",
                "铜冶炼", "铝冶炼", "铅冶炼", "锌冶炼", "镍冶炼", "锡冶炼",
                "稀土冶炼", "贵金属冶炼", "有色金属冶炼", "黑色金属冶炼",
                "合金制造", "金属加工", "金属制品", "五金", "金属结构"
            ],

            "电镀与表面处理": [
                "电镀", "镀锌", "镀铬", "镀镍", "镀铜", "镀锡", "镀金", "镀银",
                "阳极氧化", "化学镀", "磷化", "钝化", "发黑", "表面处理",
                "喷涂", "喷漆", "涂装", "烤漆", "电泳", "金属防腐", "防锈处理"
            ],

            # 化工行业（多种重金属和有机物）
            "基础化学原料": [
                "化工厂", "化学原料", "基础化工", "无机化工", "有机化工",
                "酸类生产", "碱类生产", "盐类生产", "化学试剂", "催化剂",
                "硫酸", "盐酸", "硝酸", "磷酸", "氢氟酸", "烧碱", "纯碱",
                "铬盐", "锰盐", "铅盐", "锌盐", "铜盐", "镍盐", "钴盐"
            ],

            "肥料与农药制造": [
                "化肥厂", "农药厂", "肥料生产", "农药生产", "杀虫剂", "除草剂",
                "杀菌剂", "植物生长剂", "磷肥", "氮肥", "钾肥", "复合肥",
                "有机肥", "生物农药", "化学农药", "农用化学品"
            ],

            "涂料油墨颜料": [
                "涂料厂", "油漆厂", "油墨厂", "颜料厂", "染料厂", "色浆厂",
                "钛白粉", "氧化铁", "铬黄", "镉红", "钴蓝", "铅铬黄", "锌钡白",
                "着色剂", "填充剂", "助剂", "树脂生产"
            ],

            "合成材料制造": [
                "塑料厂", "橡胶厂", "化纤厂", "合成材料", "高分子材料",
                "聚氯乙烯", "聚丙烯", "聚乙烯", "聚酯", "尼龙", "腈纶",
                "丁苯橡胶", "顺丁橡胶", "氯丁橡胶", "硅橡胶"
            ],

            "专用化学品": [
                "专用化学品", "电子化学品", "精细化工", "医药中间体",
                "添加剂", "助剂", "催化剂", "表面活性剂", "水处理剂",
                "印染助剂", "皮革助剂", "造纸助剂", "油田化学品"
            ],

            # 电子电气行业（重金属和有机污染物）
            "电子元器件": [
                "电子厂", "电路板", "PCB", "半导体", "集成电路", "芯片",
                "电子元件", "电子器件", "显示器", "液晶", "LED", "光伏",
                "电池", "蓄电池", "锂电池", "铅酸电池", "镍镉电池",
                "电镀线路板", "蚀刻", "显影", "封装", "测试"
            ],

            "电气机械": [
                "电机", "发电机", "电动机", "变压器", "整流器", "电容器",
                "开关", "继电器", "电缆", "电线", "绝缘材料", "电工器材"
            ],

            # 石油化工行业
            "石油开采加工": [
                "石油开采", "天然气开采", "油气田", "采油", "采气", "钻井",
                "炼油", "石油化工", "石化", "油气加工", "芳烃", "烯烃",
                "催化裂化", "重整", "加氢", "脱硫", "油气储运"
            ],

            "焦化与核燃料": [
                "焦化", "焦炭", "煤化工", "煤气化", "煤液化", "煤制气",
                "煤制油", "核燃料", "铀浓缩", "核材料", "放射性"
            ],

            # 纺织皮革行业
            "纺织印染": [
                "纺织", "纺纱", "织布", "印染", "染整", "印花", "染色",
                "浆纱", "退浆", "煮炼", "漂白", "丝光", "整理", "涂层",
                "纺织助剂", "染料生产", "纺织化工"
            ],

            "皮革毛皮": [
                "皮革", "制革", "毛皮", "鞣制", "硝制", "皮革加工",
                "皮草", "裘皮", "皮革制品", "皮鞋", "皮具"
            ],

            # 造纸与印刷行业
            "造纸业": [
                "造纸", "纸浆", "纸制品", "纸业", "制浆", "漂白纸浆",
                "废纸造纸", "文化用纸", "包装用纸", "特种纸", "纸板"
            ],

            "印刷业": [
                "印刷", "彩印", "胶印", "凹印", "丝印", "柔印", "数码印刷",
                "印务", "印刷包装", "印刷材料", "油墨生产"
            ],

            # 非金属矿物制品
            "水泥与建材": [
                "水泥", "水泥厂", "混凝土", "预拌混凝土", "砂浆", "建材",
                "石灰", "石膏", "砖瓦", "陶瓷", "瓷砖", "玻璃", "玻璃纤维",
                "耐火材料", "保温材料", "石材加工", "石料", "砂石"
            ],

            # 废物处理与资源化
            "废物处理": [
                "废物处理", "危险废物", "医疗废物", "固废处理", "危废处置",
                "垃圾焚烧", "填埋场", "渗滤液", "废物利用", "资源回收",
                "废金属回收", "电子废物", "废电池", "废油处理", "废酸处理"
            ],

            # 其他重污染行业
            "机械制造": [
                "机械制造", "设备制造", "重型机械", "工程机械", "机床",
                "铸造", "锻造", "热处理", "焊接", "切割", "机加工"
            ],

            "交通运输设备": [
                "汽车制造", "汽车配件", "摩托车制造", "船舶制造", "修船",
                "拆船", "飞机制造", "铁路设备", "集装箱制造"
            ],

            "仓储与物流": [
                "仓储", "物流园", "货场", "堆场", "仓库", "储运", "货运站",
                "油库", "气库", "化学品仓库", "危险品仓库"
            ],

            # 食品与医药行业
            "食品加工": [
                "食品厂", "屠宰场", "肉类加工", "水产加工", "酿造", "发酵",
                "味精", "酱油", "醋", "食品添加剂", "饲料加工"
            ],

            "医药制造": [
                "制药", "药厂", "原料药", "医药中间体", "生物制药",
                "化学制药", "中药提取", "制剂", "医疗器械"
            ],

            # 新增：土壤重金属重点监管行业
            # 其他潜在污染行业
            "其他重工业": ["机械制造", "设备制造", "重型机械", "工业园", "工业园区", "制造厂"]

        }

    def wgs84_to_gcj02(self, lng, lat):
        """
        WGS84转GCJ02坐标系（火星坐标系）
        使用公开的转换算法
        """
        pi = 3.1415926535897932384626
        a = 6378245.0
        ee = 0.00669342162296594323

        def transform_lng(lng, lat):
            ret = 300.0 + lng + 2.0 * lat + 0.1 * lng * lng + 0.1 * lng * lat + 0.1 * math.sqrt(abs(lng))
            ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 * math.sin(2.0 * lng * pi)) * 2.0 / 3.0
            ret += (20.0 * math.sin(lng * pi) + 40.0 * math.sin(lng / 3.0 * pi)) * 2.0 / 3.0
            ret += (150.0 * math.sin(lng / 12.0 * pi) + 300.0 * math.sin(lng / 30.0 * pi)) * 2.0 / 3.0
            return ret

        def transform_lat(lng, lat):
            ret = -100.0 + 2.0 * lng + 3.0 * lat + 0.2 * lat * lat + 0.1 * lng * lat + 0.2 * math.sqrt(abs(lng))
            ret += (20.0 * math.sin(6.0 * lng * pi) + 20.0 * math.sin(2.0 * lng * pi)) * 2.0 / 3.0
            ret += (20.0 * math.sin(lat * pi) + 40.0 * math.sin(lat / 3.0 * pi)) * 2.0 / 3.0
            ret += (160.0 * math.sin(lat / 12.0 * pi) + 320.0 * math.sin(lat * pi / 30.0)) * 2.0 / 3.0
            return ret

        dlat = transform_lat(lng - 105.0, lat - 35.0)
        dlng = transform_lng(lng - 105.0, lat - 35.0)

        radlat = lat / 180.0 * pi
        magic = math.sin(radlat)
        magic = 1 - ee * magic * magic
        sqrtmagic = math.sqrt(magic)

        dlat = (dlat * 180.0) / ((a * (1 - ee)) / (magic * sqrtmagic) * pi)
        dlng = (dlng * 180.0) / (a / sqrtmagic * math.cos(radlat) * pi)

        mglat = lat + dlat
        mglng = lng + dlng

        return mglng, mglat

    def calculate_distance(self, lng1, lat1, lng2, lat2):
        """使用Haversine公式计算两点间距离（米）"""
        lng1, lat1, lng2, lat2 = map(math.radians, [lng1, lat1, lng2, lat2])

        dlng = lng2 - lng1
        dlat = lat2 - lat1
        a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlng / 2) ** 2
        c = 2 * math.asin(math.sqrt(a))
        r = 6371
        return c * r * 1000

    def is_pollution_enterprise(self, name, enterprise_type):
        """判断是否为污染源企业"""
        combined_text = (name + ' ' + enterprise_type).lower()

        for industry, keywords in self.pollution_industry_types.items():
            for keyword in keywords:
                if keyword.lower() in combined_text:
                    return True, industry

        return False, "其他"

    def search_nearby_pollution_sources(self, lng, lat, radius=5000):
        """搜索周边污染源企业"""
        gcj_lng, gcj_lat = self.wgs84_to_gcj02(lng, lat)

        print(f"  坐标转换: WGS84({lng:.6f}, {lat:.6f}) -> GCJ02({gcj_lng:.6f}, {gcj_lat:.6f})")

        all_enterprises = []

        type_codes = [
            "060000", "060100", "060200", "060300", "060400",
            "070000", "080000", "170000", "010000", "020000",
            "030000", "180000", "190000",
        ]

        for type_code in type_codes:
            page = 1
            while True:
                params = {
                    'key': self.api_key,
                    'location': f'{gcj_lng},{gcj_lat}',
                    'radius': radius,
                    'types': type_code,
                    'offset': 50,
                    'page': page,
                    'extensions': 'all'
                }

                try:
                    response = requests.get(self.base_url, params=params, timeout=10)
                    data = response.json()

                    # 检查API密钥是否有效
                    if data.get('status') == '0':
                        error_info = data.get('info', '')
                        if 'key' in error_info.lower() or '配额' in error_info:
                            raise Exception(f"API密钥无效或配额已用完: {error_info}")

                    if data['status'] == '1' and int(data.get('count', 0)) > 0:
                        pois = data['pois']

                        for poi in pois:
                            if 'location' in poi and poi['location']:
                                try:
                                    enterprise_lng, enterprise_lat = map(float, poi['location'].split(','))
                                except:
                                    continue

                                distance = self.calculate_distance(lng, lat, enterprise_lng, enterprise_lat)

                                is_pollution, industry_type = self.is_pollution_enterprise(
                                    poi.get('name', ''),
                                    poi.get('type', '')
                                )

                                if is_pollution:
                                    enterprise_info = {
                                        'enterprise_name': poi.get('name', ''),
                                        'enterprise_type': poi.get('type', ''),
                                        'enterprise_address': poi.get('address', ''),
                                        'enterprise_longitude': enterprise_lng,
                                        'enterprise_latitude': enterprise_lat,
                                        'distance_to_sample': round(distance),
                                        'pollution_industry': industry_type,
                                        'business_area': poi.get('business_area', ''),
                                        'telephone': poi.get('tel', ''),
                                        'poi_id': poi.get('id', '')
                                    }
                                    all_enterprises.append(enterprise_info)

                        if len(pois) < 50:
                            break
                        page += 1

                        time.sleep(0.2)
                    else:
                        break

                except Exception as e:
                    print(f"  API请求错误: {e}")
                    # 如果是API密钥问题，直接抛出异常
                    if 'key' in str(e).lower() or '配额' in str(e):
                        raise e
                    break

        return all_enterprises


def load_sample_data(csv_file_path):
    """从CSV文件加载采样点数据"""
    try:
        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-8-sig']
        sample_data = None

        for encoding in encodings:
            try:
                sample_data = pd.read_csv(csv_file_path, encoding=encoding)
                print(f"成功使用 {encoding} 编码读取文件")
                break
            except UnicodeDecodeError:
                continue

        if sample_data is None:
            print("无法读取CSV文件，请检查文件路径和编码")
            return None

        required_columns = ['序号', '经度', '纬度']
        missing_columns = [col for col in required_columns if col not in sample_data.columns]

        if missing_columns:
            print(f"CSV文件缺少必要的列: {missing_columns}")
            print(f"文件中的列: {list(sample_data.columns)}")
            return None

        sample_data = sample_data.rename(columns={
            '序号': 'sample_id',
            '经度': 'longitude',
            '纬度': 'latitude'
        })

        print(f"成功加载 {len(sample_data)} 个采样点")
        return sample_data

    except Exception as e:
        print(f"加载CSV文件时出错: {e}")
        return None


class ExcelWriter:
    """Excel写入器，支持实时追加数据"""

    def __init__(self, output_file):
        self.output_file = output_file
        self.initialized = False

    def initialize_excel(self, columns):
        """初始化Excel文件"""
        df = pd.DataFrame(columns=columns)
        df.to_excel(self.output_file, index=False)
        self.initialized = True
        print(f"已创建Excel文件: {self.output_file}")

    def append_to_excel(self, data_list):
        """追加数据到Excel文件"""
        if not data_list:
            return

        if not self.initialized:
            # 使用第一条数据的键作为列名
            columns = list(data_list[0].keys())
            self.initialize_excel(columns)

        try:
            # 读取现有数据
            if os.path.exists(self.output_file):
                existing_df = pd.read_excel(self.output_file)
            else:
                existing_df = pd.DataFrame()

            # 创建新数据DataFrame
            new_df = pd.DataFrame(data_list)

            # 合并数据
            if not existing_df.empty:
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            else:
                combined_df = new_df

            # 保存到Excel
            combined_df.to_excel(self.output_file, index=False)

            print(f"  已写入 {len(data_list)} 条数据到Excel，当前总计 {len(combined_df)} 条")

        except Exception as e:
            print(f"  写入Excel时出错: {e}")

    def format_excel(self):
        """格式化Excel文件"""
        try:
            if not os.path.exists(self.output_file):
                return

            workbook = load_workbook(self.output_file)
            worksheet = workbook.active

            # 设置标题样式
            header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
            header_font = Font(bold=True)

            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 设置列宽
            column_widths = {
                'A': 12, 'B': 15, 'C': 15, 'D': 30, 'E': 30,
                'F': 40, 'G': 15, 'H': 15, 'I': 15, 'J': 20,
                'K': 15, 'L': 20
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            workbook.save(self.output_file)
            print(f"已格式化Excel文件: {self.output_file}")

        except Exception as e:
            print(f"格式化Excel时出错: {e}")


def collect_pollution_sources(sample_data, api_key, output_file):
    """
    为每个采样点收集周边污染源企业信息，实时写入Excel
    """
    collector = AccuratePollutionSourceCollector(api_key)
    excel_writer = ExcelWriter(output_file)

    total_sources_found = 0
    processed_samples = 0

    print("开始收集污染源企业信息...")

    for index, sample in sample_data.iterrows():
        processed_samples += 1
        print(f"处理采样点 {processed_samples}/{len(sample_data)}: {sample['sample_id']}")

        try:
            enterprises = collector.search_nearby_pollution_sources(
                sample['longitude'],
                sample['latitude'],
                radius=5000
            )

            if enterprises:
                # 准备要写入的数据
                data_to_write = []
                for enterprise in enterprises:
                    result = {
                        'sample_id': sample['sample_id'],
                        'sample_longitude': sample['longitude'],
                        'sample_latitude': sample['latitude'],
                        **{f'sample_{col}': sample.get(col, '') for col in sample_data.columns
                           if col not in ['sample_id', 'longitude', 'latitude']},
                        **enterprise
                    }
                    data_to_write.append(result)

                # 实时写入Excel
                excel_writer.append_to_excel(data_to_write)
                total_sources_found += len(enterprises)
                print(f"  找到 {len(enterprises)} 个潜在污染源，已实时保存")
            else:
                print(f"  未找到污染源")

        except Exception as e:
            error_msg = str(e)
            if 'key' in error_msg.lower() or '配额' in error_msg:
                print(f"\n API密钥错误或配额已用完: {error_msg}")
                print("正在保存已收集的数据...")

                # 格式化最终的Excel文件
                excel_writer.format_excel()

                # 显示统计信息
                if total_sources_found > 0:
                    print(f"\n已成功收集 {total_sources_found} 个污染源信息")
                    print(f"已处理 {processed_samples}/{len(sample_data)} 个采样点")
                    print(f"数据已保存到: {output_file}")
                else:
                    print("未收集到任何污染源信息")

                return False, total_sources_found, processed_samples
            else:
                print(f"  处理采样点 {sample['sample_id']} 时出错: {e}")

        # 采样点间延迟，避免API限制
        time.sleep(0.5)

    # 所有采样点处理完成，格式化Excel
    excel_writer.format_excel()
    return True, total_sources_found, processed_samples


def main():
    # 您的高德地图API密钥
    API_KEY = "baaa323856298f0ddc39f0179d8fca4a"

    # CSV文件路径
    csv_file_path = "采样点数据.csv"

    # 输出文件路径
    output_file = '污染源企业信息.xlsx'

    # 检查文件是否存在
    if not os.path.exists(csv_file_path):
        print(f"文件不存在: {csv_file_path}")
        return

    # 加载采样点数据
    sample_data = load_sample_data(csv_file_path)
    if sample_data is None:
        return

    print("开始执行污染源企业信息收集...")
    print(f"结果将实时保存到: {output_file}")

    # 收集污染源企业信息（实时写入Excel）
    success, total_sources, processed_samples = collect_pollution_sources(
        sample_data, API_KEY, output_file
    )

    if success:
        print(f"\n 收集完成！")
        print(f"处理采样点数量: {processed_samples}/{len(sample_data)}")
        print(f"找到污染源总数: {total_sources}")
        print(f"结果已保存至: {output_file}")

        # 如果收集到了数据，显示统计信息
        if total_sources > 0:
            try:
                df = pd.read_excel(output_file)
                industry_stats = df['pollution_industry'].value_counts()
                print("\n污染源行业分布:")
                for industry, count in industry_stats.items():
                    print(f"  {industry}: {count}个")

            except Exception as e:
                print(f"读取统计信息时出错: {e}")
    else:
        print(f"\n收集过程中断！")
        print(f"已处理采样点: {processed_samples}/{len(sample_data)}")
        print(f"已找到污染源: {total_sources}个")
        print(f"部分结果已保存至: {output_file}")
        print("请检查API密钥或等待配额重置后继续执行")


if __name__ == "__main__":
    main()